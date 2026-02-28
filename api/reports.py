# -*- coding: utf-8 -*-
"""报表导出 - Excel 薪资报表"""
import io
from datetime import datetime
from flask import Blueprint, request, send_file, jsonify
from flask_jwt_extended import jwt_required

from models import get_db, dict_from_row, date_month_sql
from api.auth import require_admin, get_current_employee_id

reports_bp = Blueprint('reports', __name__)

@reports_bp.route('/salary-excel', methods=['GET'])
@jwt_required()
def export_salary_excel():
    """导出 Excel 薪资报表 - 用于发薪、对账"""
    emp_id = get_current_employee_id()
    employee_id = request.args.get('employee_id', type=int)
    month = request.args.get('month', '').strip()
    if not month:
        return jsonify({'ok': False, 'msg': '请指定月份'}), 400
    if emp_id and employee_id and employee_id != emp_id:
        return jsonify({'ok': False, 'msg': '无权限'}), 403
    if emp_id:
        employee_id = emp_id
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side
    except ImportError:
        return jsonify({'ok': False, 'msg': '请安装 openpyxl'}), 500
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'{month}薪资报表'
    header_font = Font(bold=True)
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = ['日期', '员工姓名', '工号', '款式', '尺码', '工序', '数量', '单价', '金额', '工单号']
    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        c.font = header_font
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')
    with get_db() as conn:
        c = conn.cursor()
        sql = f"""
            SELECT pr.record_date, e.name as employee_name, e.employee_no, pr.style_code, pr.size, pr.process_name,
                   pr.quantity, pr.unit_price, pr.amount, wo.order_no
            FROM piece_records pr
            JOIN employees e ON pr.employee_id = e.id
            LEFT JOIN work_orders wo ON pr.work_order_id = wo.id
            WHERE {date_month_sql('pr.record_date')} = ?
        """
        params = [month]
        if employee_id:
            sql += " AND pr.employee_id = ?"
            params.append(employee_id)
        sql += " ORDER BY e.name, pr.record_date, pr.style_code"
        c.execute(sql, params)
        rows = c.fetchall()
    for row_idx, row in enumerate(rows, 2):
        r = dict_from_row(row)
        ws.cell(row_idx, 1, r.get('record_date', ''))
        ws.cell(row_idx, 2, r.get('employee_name', ''))
        ws.cell(row_idx, 3, r.get('employee_no', '') or '')
        ws.cell(row_idx, 4, r.get('style_code', ''))
        ws.cell(row_idx, 5, r.get('size', '') or '')
        ws.cell(row_idx, 6, r.get('process_name', '') or '')
        ws.cell(row_idx, 7, r.get('quantity', 0))
        ws.cell(row_idx, 8, r.get('unit_price', 0))
        ws.cell(row_idx, 9, r.get('amount', 0))
        ws.cell(row_idx, 10, r.get('order_no', '') or '')
        for col in range(1, 11):
            ws.cell(row_idx, col).border = border
    # 汇总行
    total_row = len(rows) + 2
    ws.cell(total_row, 1, '合计')
    ws.cell(total_row, 7, sum(r.get('quantity', 0) for r in [dict_from_row(x) for x in rows]))
    ws.cell(total_row, 9, sum(r.get('amount', 0) for r in [dict_from_row(x) for x in rows]))
    for col in range(1, 11):
        ws.cell(total_row, col).font = header_font
        ws.cell(total_row, col).border = border
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f'薪资报表_{month}_{datetime.now().strftime("%Y%m%d%H%M")}.xlsx'
    return send_file(buf, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
